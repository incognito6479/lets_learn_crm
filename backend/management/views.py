from rest_framework import viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django.utils import timezone
from django.db.models import Sum, Q
from rest_framework.exceptions import ValidationError
from decimal import Decimal
import datetime
import re
import openpyxl
from .models import Enrollment
from .models import Branch, User, Student, Room, Course, Group, Enrollment, Payment, Grade, Absence
from .serializers import (
    BranchSerializer, UserSerializer, StudentSerializer, 
    RoomSerializer, CourseSerializer, GroupSerializer, 
    EnrollmentSerializer, PaymentSerializer, GradeSerializer, AbsenceSerializer
)
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView

class SoftDeleteModelViewSet(viewsets.ModelViewSet):
    def get_queryset(self):
        return self.queryset.filter(is_active=True)

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.save()


class IsAdminOrCEOOrSuperuserOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.role in ['admin', 'CEO', 'superuser'] or request.user.is_superuser

class IsAdminOrCEOOrSuperuserOrSelf(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.role in ['admin', 'CEO', 'superuser'] or request.user.is_superuser

    def has_object_permission(self, request, view, obj):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.user.id == obj.id:
            return True
        return request.user.role in ['admin', 'CEO', 'superuser'] or request.user.is_superuser

class IsCashierOrAdminOrCEOOrSuperuserOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.role in ['admin', 'CEO', 'superuser', 'cashier'] or request.user.is_superuser

class IsCashierOrAdminOrCEOOrSuperuser(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        return request.user.role in ['admin', 'CEO', 'superuser', 'cashier'] or request.user.is_superuser

class IsTeacherOrAdminOrCEOOrSuperuser(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        return request.user.role in ['admin', 'CEO', 'superuser', 'teacher'] or request.user.is_superuser

class BranchViewSet(SoftDeleteModelViewSet):
    queryset = Branch.objects.all()
    serializer_class = BranchSerializer
    permission_classes = [IsAdminOrCEOOrSuperuserOrReadOnly]

class UserViewSet(SoftDeleteModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdminOrCEOOrSuperuserOrSelf]

    def get_queryset(self):
        qs = self.queryset.all()
        user = self.request.user
        if user.is_authenticated and user.role == 'admin' and user.branch:
            qs = qs.filter(Q(branch=user.branch) | Q(branch__isnull=True) | Q(group__branch=user.branch)).distinct()
        return qs

class StudentViewSet(SoftDeleteModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    permission_classes = [IsCashierOrAdminOrCEOOrSuperuserOrReadOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == 'admin' and user.branch:
            qs = qs.filter(
                Q(enrollment__group__branch=user.branch) | 
                Q(enrollment__isnull=True) | 
                Q(enrollment__group__branch__isnull=True)
            ).distinct()
        return qs

class RoomViewSet(SoftDeleteModelViewSet):
    queryset = Room.objects.all()
    serializer_class = RoomSerializer
    permission_classes = [IsAdminOrCEOOrSuperuserOrReadOnly]

class CourseViewSet(SoftDeleteModelViewSet):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [IsAdminOrCEOOrSuperuserOrReadOnly]

class GroupViewSet(SoftDeleteModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [IsAdminOrCEOOrSuperuserOrReadOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == 'admin' and user.branch:
            qs = qs.filter(Q(branch=user.branch) | Q(branch__isnull=True))
        return qs

    def perform_update(self, serializer):
        instance = self.get_object()
        new_status = serializer.validated_data.get('status', instance.status)
        
        if new_status == 'finished' and instance.status != 'finished':
            has_debt = Enrollment.objects.filter(
                group=instance,
                status='enrolled',
                payment_status='debt'
            ).exists()
            if has_debt:
                raise ValidationError("Cannot finish group because some enrolled students have outstanding debt.")
        
        group = serializer.save()
        
        if group.status == 'finished' and instance.status != 'finished':
            Enrollment.objects.filter(group=group, status='enrolled').update(status='finished')

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        excel_file = request.FILES.get('file')
        if not excel_file:
            raise ValidationError("No file was uploaded.")

        # Resolve teacher name from filename
        filename = excel_file.name
        name_part = filename.rsplit('.', 1)[0]
        teacher_name = name_part.strip().replace('_', ' ').replace('-', ' ')
        
        # Get or create teacher
        username_clean = re.sub(r'[^a-zA-Z0-9_.-]', '', name_part.strip())
        if not username_clean:
            username_clean = f"teacher_{int(timezone.now().timestamp())}"
        
        username_clean = username_clean.lower()
        
        teacher, created = User.objects.get_or_create(
            username=username_clean,
            defaults={
                'first_name': teacher_name,
                'role': 'teacher',
                'is_active': True
            }
        )
        if created:
            teacher.set_password('teacher123')
            teacher.save()

        # Load workbook using openpyxl
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
        except Exception as e:
            raise ValidationError(f"Error opening Excel file: {str(e)}")

        imported_groups = []
        total_students_imported = 0
        total_payments_imported = 0
        total_absences_imported = 0

        try:
            current_month = int(request.data.get('month', timezone.localdate().month))
        except (ValueError, TypeError):
            current_month = timezone.localdate().month

        try:
            current_year = int(request.data.get('year', timezone.localdate().year))
        except (ValueError, TypeError):
            current_year = timezone.localdate().year

        try:
            group_price = Decimal(str(request.data.get('price', '450000.00')).replace(' ', '').replace(',', ''))
        except (ValueError, TypeError, Decimal.InvalidOperation):
            group_price = Decimal('450000.00')

        for sheet in wb.worksheets:
            sheet_name = sheet.title
            name_clean = sheet_name.replace(';', ':').replace(' ', '').lower()
            
            # Extract start time
            starts_at = "09:00:00"
            colon_match = re.search(r'(\d+)[:;](\d+)', sheet_name)
            if colon_match:
                hour = int(colon_match.group(1))
                minute = int(colon_match.group(2))
                starts_at = f"{hour:02d}:{minute:02d}:00"
            else:
                time_match = re.search(r'\d+', name_clean)
                if time_match:
                    num_str = time_match.group()
                    if len(num_str) == 3:
                        starts_at = f"0{num_str[0]}:{num_str[1:]}:00"
                    elif len(num_str) == 4:
                        starts_at = f"{num_str[:2]}:{num_str[2:]}:00"
                    elif len(num_str) in [1, 2]:
                        hour = int(num_str)
                        starts_at = f"{hour:02d}:00:00"

            # Extract days
            days = 'Mon-Wed-Fri'
            if any(d in name_clean for d in ['tue', 'thu', 'sat']):
                days = 'Tue-Thur-Sat'

            # Create group name
            group_name = f"{teacher_name} - {sheet_name}"
            
            # Create Group
            group = Group.objects.create(
                name=group_name,
                teacher=teacher,
                starts_at=starts_at,
                duration=90, # 1.5 h = 90 mins
                group_days_at=days,
                price=group_price,
                started_at=datetime.date(current_year, current_month, 1),
                status='ongoing'
            )
            imported_groups.append(group_name)

            # Helper for date parsing
            def parse_enrollment_date(date_val):
                if not date_val:
                    return datetime.date(current_year, current_month, 1)
                if isinstance(date_val, (datetime.date, datetime.datetime)):
                    return date_val.date() if isinstance(date_val, datetime.datetime) else date_val
                
                date_str = str(date_val).lower().strip()
                match = re.search(r'(\d+)\s*([a-zа-я]+|\/\s*[a-z]+)', date_str)
                if not match:
                    return datetime.date(current_year, current_month, 1)
                
                day = int(match.group(1))
                month_part = match.group(2).replace('/', '').strip()
                
                month_map = {
                    'yan': 1, 'yanvar': 1, 'jan': 1,
                    'fev': 2, 'fevral': 2, 'feb': 2,
                    'mar': 3, 'mart': 3,
                    'apr': 4, 'aprel': 4,
                    'may': 5,
                    'iyun': 6, 'jun': 6,
                    'iyul': 7, 'jul': 7,
                    'avg': 8, 'avgust': 8, 'aug': 8,
                    'sen': 9, 'sentyabr': 9, 'sep': 9,
                    'okt': 10, 'oktyabr': 10, 'oct': 10,
                    'noy': 11, 'noyabr': 11, 'nov': 11,
                    'dek': 12, 'dekabr': 12, 'dec': 12
                }
                
                month = month_map.get(month_part, current_month)
                try:
                    return datetime.date(current_year, month, day)
                except ValueError:
                    return datetime.date(current_year, current_month, 1)

            # Loop through student rows starting from row 3
            for r_idx in range(3, sheet.max_row + 1):
                name_val = sheet.cell(row=r_idx, column=2).value
                if not name_val or str(name_val).strip() == "" or str(name_val).strip().upper() == "ОБЩИЙ ИТОГ":
                    break
                
                student_name = str(name_val).strip()
                
                enroll_date_val = sheet.cell(row=r_idx, column=3).value
                enrollment_date = parse_enrollment_date(enroll_date_val)
                
                phone_val = sheet.cell(row=r_idx, column=37).value # Col AK
                phone_str = str(phone_val or '').replace(' ', '').replace('-', '').strip()
                if phone_str and not phone_str.startswith('+') and len(phone_str) == 9:
                    phone_str = f"+998{phone_str}"
                elif not phone_str:
                    phone_str = "+998900000000"
                phone_str = phone_str[:20]

                # Find or create Student
                student, _ = Student.objects.get_or_create(
                    full_name=student_name,
                    defaults={'phone1': phone_str}
                )
                total_students_imported += 1

                # Create Enrollment
                enrollment = Enrollment.objects.create(
                    student=student,
                    group=group,
                    date=enrollment_date,
                    status='enrolled'
                )

                # Auto-generate payments for prior months to mark them as paid
                temp_year = enrollment_date.year
                temp_month = enrollment_date.month
                target_year = current_year
                target_month = current_month

                while (temp_year < target_year) or (temp_year == target_year and temp_month < target_month):
                    pay_day = min(enrollment_date.day, 28)
                    pay_date = datetime.date(temp_year, temp_month, pay_day)
                    
                    Payment.objects.create(
                        group=group,
                        student=student,
                        amount=group_price,
                        payment_method='cash',
                        status='accepted',
                        description=f"Auto-generated payment for prior month: {pay_date.strftime('%B %Y')}",
                        payment_date=timezone.make_aware(datetime.datetime.combine(pay_date, datetime.time(12, 0)))
                    )
                    total_payments_imported += 1
                    
                    # Increment month
                    if temp_month == 12:
                        temp_month = 1
                        temp_year += 1
                    else:
                        temp_month += 1

                # Col AI (35): Total paid
                paid_val = sheet.cell(row=r_idx, column=35).value
                try:
                    total_paid = float(paid_val or 0) * 1000
                except (ValueError, TypeError):
                    total_paid = 0.0

                if total_paid > 0:
                    Payment.objects.create(
                        group=group,
                        student=student,
                        amount=Decimal(str(total_paid)),
                        payment_method='cash',
                        status='accepted',
                        description="Imported total paid from Excel sheet"
                    )
                    total_payments_imported += 1

                # Col AJ (36): Debt / Overpaid
                note_cell = sheet.cell(row=r_idx, column=36)
                note_val = note_cell.value
                
                # Check color
                color_hex = None
                if note_cell.font and note_cell.font.color and note_cell.font.color.rgb:
                    color_hex = note_cell.font.color.rgb
                if not color_hex and note_cell.fill and note_cell.fill.start_color and note_cell.fill.start_color.rgb:
                    color_hex = note_cell.fill.start_color.rgb
                
                is_green = False
                if color_hex and isinstance(color_hex, str):
                    color_clean = color_hex.upper().strip()
                    if color_clean.startswith('FF'):
                        color_clean = color_clean[2:]
                    if len(color_clean) == 6:
                        red_val = int(color_clean[0:2], 16)
                        green_val = int(color_clean[2:4], 16)
                        blue_val = int(color_clean[4:6], 16)
                        if green_val > 120 and red_val < 100 and blue_val < 100:
                            is_green = True

                if is_green and note_val:
                    try:
                        overpaid_val = float(note_val) * 1000
                    except (ValueError, TypeError):
                        overpaid_val = 0.0
                    
                    if overpaid_val > 0:
                        Payment.objects.create(
                            group=group,
                            student=student,
                            amount=Decimal(str(overpaid_val)),
                            payment_method='cash',
                            status='accepted',
                            description="Imported overpayment from Excel sheet"
                        )
                        total_payments_imported += 1

                # Recalculate enrollment debt
                enrollment.check_debt()

                # Col D (4) to AH (34): Absences
                for day_col in range(4, 35):
                    day_num = day_col - 3
                    cell_val = sheet.cell(row=r_idx, column=day_col).value
                    if cell_val and str(cell_val).strip() == "-":
                        Absence.objects.create(
                            student=student,
                            group=group,
                            teacher=teacher,
                            date=datetime.date(current_year, current_month, day_num)
                        )
                        total_absences_imported += 1

        return Response({
            'status': 'success',
            'imported_groups': imported_groups,
            'total_students': total_students_imported,
            'total_payments': total_payments_imported,
            'total_absences': total_absences_imported
        })

class EnrollmentViewSet(SoftDeleteModelViewSet):
    queryset = Enrollment.objects.all()
    serializer_class = EnrollmentSerializer
    permission_classes = [IsCashierOrAdminOrCEOOrSuperuserOrReadOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == 'admin' and user.branch:
            qs = qs.filter(Q(group__branch=user.branch) | Q(group__branch__isnull=True))
        return qs

    def perform_destroy(self, instance):
        instance.status = 'dropped'
        instance.save()

class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = [IsCashierOrAdminOrCEOOrSuperuser]

    def get_queryset(self):
        qs = self.queryset.all()
        user = self.request.user
        if user.is_authenticated and user.role == 'admin' and user.branch:
            qs = qs.filter(Q(group__branch=user.branch) | Q(group__branch__isnull=True))
        return qs

    def perform_create(self, serializer):
        payment = serializer.save()
        if payment.student:
            enrollment = Enrollment.objects.filter(
                student=payment.student,
                group=payment.group,
                is_active=True
            ).first()
            if enrollment:
                enrollment.check_debt()

    def perform_destroy(self, instance):
        instance.is_active = False
        instance.status = 'canceled'
        instance.save()

class GradeViewSet(SoftDeleteModelViewSet):
    queryset = Grade.objects.all()
    serializer_class = GradeSerializer
    permission_classes = [IsTeacherOrAdminOrCEOOrSuperuser]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == 'admin' and user.branch:
            qs = qs.filter(Q(group__branch=user.branch) | Q(group__branch__isnull=True))
        return qs

class AbsenceViewSet(SoftDeleteModelViewSet):
    queryset = Absence.objects.all()
    serializer_class = AbsenceSerializer
    permission_classes = [IsTeacherOrAdminOrCEOOrSuperuser]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == 'admin' and user.branch:
            qs = qs.filter(Q(group__branch=user.branch) | Q(group__branch__isnull=True))
        return qs

class CustomTokenObtainPairSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token['username'] = user.username
        token['role'] = user.role
        token['user_id'] = user.id
        return token

    def validate(self, attrs):
        data = super().validate(attrs)
        data['role'] = self.user.role
        data['user_id'] = self.user.id
        data['username'] = self.user.username
        return data

class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
