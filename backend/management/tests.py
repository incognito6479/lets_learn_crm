from django.test import TestCase
from decimal import Decimal
from django.utils import timezone
from rest_framework.exceptions import ValidationError
from rest_framework.test import APIRequestFactory, force_authenticate, APITestCase
from management.models import Branch, User, Student, Room, Course, Group, Enrollment, Payment, Grade, Absence, Notification
from management.views import GroupViewSet
from django.core.management import call_command
import io
import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl.styles import Font, PatternFill


class GroupFinishTest(TestCase):
    def setUp(self):
        self.branch = Branch.objects.create(name="Main Branch", description="Main campus center")
        self.teacher = User.objects.create(
            username="teacher1",
            first_name="John",
            last_name="Doe",
            role="teacher"
        )
        self.teacher.set_password("teacherpassword")
        self.teacher.save()
        self.admin = User.objects.create(
            username="admin1",
            first_name="Admin",
            last_name="User",
            role="admin"
        )
        self.admin.set_password("adminpassword")
        self.admin.save()
        self.student = Student.objects.create(
            full_name="Bob Smith",
            phone1="+998904445566"
        )
        self.course = Course.objects.create(
            name="English I",
            price=Decimal('100000.00')
        )
        self.room = Room.objects.create(
            name="Room 1",
            branch=self.branch
        )
        self.group = Group.objects.create(
            name="English Group A",
            course=self.course,
            teacher=self.teacher,
            room=self.room,
            branch=self.branch,
            price=Decimal('100000.00'),
            starts_at="09:00",
            duration=90,
            started_at=timezone.localdate(),
            status="ongoing"
        )
        self.enrollment = Enrollment.objects.create(
            student=self.student,
            group=self.group,
            status="enrolled"
        )
        self.factory = APIRequestFactory()

    def test_finish_group_with_debt_raises_validation_error(self):
        # The student has not paid anything, so they have debt of 100,000 UZS
        self.enrollment.check_debt()
        self.assertEqual(self.enrollment.payment_status, 'debt')

        # Try to patch status to finished
        view = GroupViewSet.as_view({'patch': 'partial_update'})
        request = self.factory.patch(
            f'/api/groups/{self.group.id}/',
            {'status': 'finished'},
            format='json'
        )
        force_authenticate(request, user=self.admin)
        
        response = view(request, pk=self.group.id)
        self.assertEqual(response.status_code, 400)
        self.assertIn("outstanding debt", str(response.data[0]))

        # Assert status remains ongoing
        self.group.refresh_from_db()
        self.assertEqual(self.group.status, 'ongoing')

    def test_finish_group_with_no_debt_succeeds(self):
        # Record payment to clear the debt
        Payment.objects.create(
            student=self.student,
            group=self.group,
            amount=Decimal('100000.00'),
            payment_method="cash",
            status="accepted"
        )
        self.enrollment.check_debt()
        self.assertEqual(self.enrollment.payment_status, 'paid')

        # Try to patch status to finished
        view = GroupViewSet.as_view({'patch': 'partial_update'})
        request = self.factory.patch(
            f'/api/groups/{self.group.id}/',
            {'status': 'finished'},
            format='json'
        )
        force_authenticate(request, user=self.admin)
        
        response = view(request, pk=self.group.id)
        self.assertEqual(response.status_code, 200)

        # Assert status is updated to finished
        self.group.refresh_from_db()
        self.assertEqual(self.group.status, 'finished')

        # Assert enrollment status is updated to finished
        self.enrollment.refresh_from_db()
        self.assertEqual(self.enrollment.status, 'finished')

    def test_anniversary_billing_cycle_transition(self):
        import datetime
        # Enrolled on March 23
        self.enrollment.date = datetime.date(2026, 3, 23)
        self.enrollment.save()

        # Paid once on March 23 (amount equal to group price: 100,000 UZS)
        Payment.objects.create(
            student=self.student,
            group=self.group,
            amount=Decimal('100000.00'),
            payment_method="cash",
            status="accepted",
            payment_date=timezone.make_aware(datetime.datetime(2026, 3, 23, 12, 0))
        )

        from unittest.mock import patch

        # Helper side_effect functions to mock timezone.localdate
        def localdate_april_23(value=None, timezone=None):
            if value is None:
                return datetime.date(2026, 4, 23)
            return value.date() if isinstance(value, datetime.datetime) else value

        def localdate_april_24(value=None, timezone=None):
            if value is None:
                return datetime.date(2026, 4, 24)
            return value.date() if isinstance(value, datetime.datetime) else value

        # 1. On April 23 (the anniversary date): they should still be marked as "paid"
        with patch('django.utils.timezone.localdate', side_effect=localdate_april_23):
            self.enrollment.check_debt()
            self.assertEqual(self.enrollment.payment_status, 'paid')
            self.assertEqual(float(self.enrollment.debt_amount), 0.00)

        # 2. On April 24 (the day after the anniversary): they should be marked as "debt" (100,000 UZS)
        with patch('django.utils.timezone.localdate', side_effect=localdate_april_24):
            self.enrollment.check_debt()
            self.assertEqual(self.enrollment.payment_status, 'debt')
            self.assertEqual(float(self.enrollment.debt_amount), 100000.00)

    def test_pdf_uploaded_debt_checking(self):
        import datetime
        from unittest.mock import patch

        # Enrolled on March 23
        self.enrollment.date = datetime.date(2026, 3, 23)
        self.enrollment.pdf_uploaded = True
        self.enrollment.save()

        # Register a payment on March 23
        p1 = Payment.objects.create(
            student=self.student,
            group=self.group,
            amount=Decimal('100000.00'),
            payment_method="cash",
            status="accepted",
            payment_date=timezone.make_aware(datetime.datetime(2026, 3, 23, 12, 0))
        )

        def localdate_may_23(value=None, timezone=None):
            if value is None:
                return datetime.date(2026, 5, 23)
            return value.date() if isinstance(value, datetime.datetime) else value

        # 1. On May 23 (with only March 23 payment), the student should be in debt for the current month
        # because the March 23 payment belongs to the ignored previous cycle.
        with patch('django.utils.timezone.localdate', side_effect=localdate_may_23):
            self.enrollment.check_debt()
            self.assertEqual(self.enrollment.payment_status, 'debt')
            self.assertEqual(float(self.enrollment.debt_amount), 100000.00)

        # 2. Register a payment on April 25 (which is within the current cycle of April 23 - May 23)
        p2 = Payment.objects.create(
            student=self.student,
            group=self.group,
            amount=Decimal('100000.00'),
            payment_method="cash",
            status="accepted",
            payment_date=timezone.make_aware(datetime.datetime(2026, 4, 25, 12, 0))
        )

        # Now they should be marked as "paid"
        with patch('django.utils.timezone.localdate', side_effect=localdate_may_23):
            self.enrollment.check_debt()
            self.assertEqual(self.enrollment.payment_status, 'paid')
            self.assertEqual(float(self.enrollment.debt_amount), 0.00)

    def test_clean_phone_number_logic(self):
        from management.helpers import clean_phone_number
        self.assertEqual(clean_phone_number("90 900 90 90"), "+998909009090")
        self.assertEqual(clean_phone_number("909009090"), "+998909009090")
        self.assertEqual(clean_phone_number("+998 90 900 90 90"), "+998909009090")
        self.assertEqual(clean_phone_number("998909009090"), "+998909009090")
        self.assertEqual(clean_phone_number(""), None)
        self.assertEqual(clean_phone_number(None), None)


class APIOperationsTest(APITestCase):
    def setUp(self):
        # Create default records for relations
        self.branch = Branch.objects.create(name="Test Branch", description="Test Branch Description")
        self.admin = User.objects.create_superuser(
            username="superuser",
            password="superpassword",
            role="superuser",
            branch=self.branch
        )
        self.teacher = User.objects.create(
            username="teacher_test",
            role="teacher",
            branch=self.branch
        )
        self.teacher.set_password("teacherpass")
        self.teacher.save()
        
        self.course = Course.objects.create(name="Maths", price=Decimal('150000.00'))
        self.room = Room.objects.create(name="Room 101", branch=self.branch)
        self.student = Student.objects.create(full_name="Alice Johnson", phone1="+998901112233")
        
        # Authenticate all API client calls by default with the superuser
        self.client.force_authenticate(user=self.admin)

    def test_create_group(self):
        response = self.client.post('/api/groups/', {
            'name': 'New Group',
            'course': self.course.id,
            'teacher': self.teacher.id,
            'room': self.room.id,
            'branch': self.branch.id,
            'price': '120000.00',
            'starts_at': '10:00:00',
            'duration': 90,
            'started_at': str(timezone.localdate()),
            'status': 'ongoing',
            'group_days_at': 'Mon-Wed-Fri'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(Group.objects.filter(name='New Group').count(), 1)

    def test_create_user(self):
        response = self.client.post('/api/users/', {
            'username': 'newuser',
            'password': 'newpassword',
            'role': 'cashier',
            'branch': self.branch.id,
            'first_name': 'New',
            'last_name': 'User'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        user = User.objects.get(username='newuser')
        self.assertEqual(user.role, 'cashier')
        self.assertTrue(user.check_password('newpassword'))

    def test_change_password(self):
        # Change password of the teacher user
        response = self.client.patch(f'/api/users/{self.teacher.id}/', {
            'password': 'changedpassword'
        }, format='json')
        self.assertEqual(response.status_code, 200)
        self.teacher.refresh_from_db()
        self.assertTrue(self.teacher.check_password('changedpassword'))

    def test_enroll_student(self):
        # Create group for enrollment
        group = Group.objects.create(
            name="Enrollment Test Group",
            course=self.course,
            teacher=self.teacher,
            room=self.room,
            branch=self.branch,
            price=Decimal('150000.00'),
            starts_at="10:00",
            duration=90,
            started_at=timezone.localdate()
        )
        response = self.client.post('/api/enrollments/', {
            'student': self.student.id,
            'group': group.id,
            'status': 'enrolled',
            'payment_status': 'debt'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(Enrollment.objects.filter(student=self.student, group=group).count(), 1)

    def test_exclude_student(self):
        # Exclude means soft deleting / marking enrollment as dropped
        group = Group.objects.create(
            name="Exclude Test Group",
            course=self.course,
            teacher=self.teacher,
            room=self.room,
            branch=self.branch,
            price=Decimal('150000.00'),
            starts_at="10:00",
            duration=90,
            started_at=timezone.localdate()
        )
        enrollment = Enrollment.objects.create(
            student=self.student,
            group=group,
            status='enrolled'
        )
        response = self.client.delete(f'/api/enrollments/{enrollment.id}/')
        self.assertEqual(response.status_code, 204) # Soft delete returns 204
        enrollment.refresh_from_db()
        self.assertEqual(enrollment.status, 'dropped')

    def test_make_payment(self):
        group = Group.objects.create(
            name="Payment Test Group",
            course=self.course,
            teacher=self.teacher,
            room=self.room,
            branch=self.branch,
            price=Decimal('150000.00'),
            starts_at="10:00",
            duration=90,
            started_at=timezone.localdate()
        )
        enrollment = Enrollment.objects.create(
            student=self.student,
            group=group,
            status='enrolled'
        )
        response = self.client.post('/api/payments/', {
            'student': self.student.id,
            'group': group.id,
            'amount': '150000.00',
            'payment_method': 'card',
            'status': 'accepted',
            'description': 'Test payment'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(Payment.objects.filter(student=self.student, group=group).count(), 1)
        
        # Payment trigger perform_create calls check_debt, clearing the debt
        enrollment.refresh_from_db()
        self.assertEqual(enrollment.payment_status, 'paid')
        self.assertEqual(float(enrollment.debt_amount), 0.0)

    def test_grading(self):
        group = Group.objects.create(
            name="Grading Test Group",
            course=self.course,
            teacher=self.teacher,
            room=self.room,
            branch=self.branch,
            price=Decimal('150000.00'),
            starts_at="10:00",
            duration=90,
            started_at=timezone.localdate()
        )
        enrollment = Enrollment.objects.create(
            student=self.student,
            group=group,
            status='enrolled'
        )
        response = self.client.post('/api/grades/', {
            'enrolled_student': self.student.id,
            'group': group.id,
            'teacher': self.teacher.id,
            'grade': 5,
            'date': str(timezone.localdate())
        }, format='json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(Grade.objects.filter(enrolled_student=self.student, group=group).count(), 1)

    def test_make_absence(self):
        group = Group.objects.create(
            name="Absence Test Group",
            course=self.course,
            teacher=self.teacher,
            room=self.room,
            branch=self.branch,
            price=Decimal('150000.00'),
            starts_at="10:00",
            duration=90,
            started_at=timezone.localdate()
        )
        enrollment = Enrollment.objects.create(
            student=self.student,
            group=group,
            status='enrolled'
        )
        response = self.client.post('/api/absences/', {
            'student': self.student.id,
            'group': group.id,
            'teacher': self.teacher.id,
            'date': str(timezone.localdate())
        }, format='json')
        self.assertEqual(response.status_code, 201)
        self.assertEqual(Absence.objects.filter(student=self.student, group=group).count(), 1)

    def test_jwt_obtain_and_refresh_tokens(self):
        user = User.objects.create(
            username="jwtuser",
            role="admin",
            branch=self.branch
        )
        user.set_password("jwtpassword")
        user.save()

        self.client.force_authenticate(user=None)

        # 1. Obtain token pair
        response = self.client.post('/api/token/', {
            'username': 'jwtuser',
            'password': 'jwtpassword'
        }, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertIn('access', response.data)
        self.assertIn('refresh', response.data)
        self.assertEqual(response.data['username'], 'jwtuser')
        self.assertEqual(response.data['role'], 'admin')

        # 2. Refresh access token
        refresh_token = response.data['refresh']
        response_refresh = self.client.post('/api/token/refresh/', {
            'refresh': refresh_token
        }, format='json')
        self.assertEqual(response_refresh.status_code, 200)
        self.assertIn('access', response_refresh.data)

    def test_teacher_blocked_from_payments(self):
        self.client.force_authenticate(user=self.teacher)
        response = self.client.get('/api/payments/')
        self.assertEqual(response.status_code, 403)

    def test_cashier_blocked_from_grading(self):
        cashier = User.objects.create(
            username="cashier_test",
            role="cashier",
            branch=self.branch
        )
        cashier.set_password("cashierpass")
        cashier.save()
        
        self.client.force_authenticate(user=cashier)
        response = self.client.get('/api/grades/')
        self.assertEqual(response.status_code, 403)

    def test_teacher_share_and_payouts_calculation(self):
        # Create a group specifically for this test
        group = Group.objects.create(
            name="Payout Group",
            course=self.course,
            teacher=self.teacher,
            room=self.room,
            branch=self.branch,
            price=Decimal('100000.00'),
            starts_at='10:00:00',
            duration=90,
            started_at=timezone.localdate(),
            status='ongoing',
            group_days_at='Mon-Wed-Fri',
            teacher_share=40
        )

        # Enroll the student in this group
        Enrollment.objects.create(
            student=self.student,
            group=group,
            status='enrolled'
        )

        Payment.objects.create(
            group=group,
            student=self.student,
            amount=Decimal('100000.00'),
            payment_method='cash',
            status='accepted'
        )

        self.client.force_authenticate(user=self.admin)
        response = self.client.get(f'/api/groups/{group.id}/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['teacher_earnings'], 40000.00)
        self.assertEqual(response.data['teacher_paid'], 0.00)
        self.assertEqual(response.data['teacher_remaining'], 40000.00)

        Payment.objects.create(
            group=group,
            student=None,
            teacher=self.teacher,
            amount=Decimal('15000.00'),
            payment_method='cash',
            status='accepted'
        )

        response = self.client.get(f'/api/groups/{group.id}/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['teacher_earnings'], 40000.00)
        self.assertEqual(response.data['teacher_paid'], 15000.00)
        self.assertEqual(response.data['teacher_remaining'], 25000.00)

    def test_excel_group_import(self):
        # Create mock excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tues 900"

        # Headers
        ws.cell(row=2, column=1, value="№")
        ws.cell(row=2, column=2, value="ФИО")
        ws.cell(row=2, column=3, value="дата оплаты")
        ws.cell(row=2, column=35, value="ИТОГ")
        ws.cell(row=2, column=36, value="примечание")
        ws.cell(row=2, column=37, value="телефон")

        # Row 3: student 1 - Amirshox, date: 24 iyun, total: 450, note: 450 (debt - red text), phone: 93 727 70 07
        ws.cell(row=3, column=1, value=1)
        ws.cell(row=3, column=2, value="Amirshox Davronov")
        ws.cell(row=3, column=3, value="24 iyun")
        ws.cell(row=3, column=4, value="-") # day 1 absence
        ws.cell(row=3, column=35, value=450)
        ws.cell(row=3, column=37, value="93 727 70 07")
        
        red_font = Font(color="FFFF0000")
        ws.cell(row=3, column=36, value=450).font = red_font

        # Apply yellow background fill to student name cell
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws.cell(row=3, column=2).fill = yellow_fill

        # Row 4: student 2 - Timur, date: 16/Jun, total: 0, note: 450 (overpaid - green text), phone: 88 287 17 72
        ws.cell(row=4, column=1, value=2)
        ws.cell(row=4, column=2, value="Timur Mustafoev")
        ws.cell(row=4, column=3, value="16/Jun")
        ws.cell(row=4, column=35, value=0)
        ws.cell(row=4, column=37, value="88 287 17 72")
        
        green_font = Font(color="FF00FF00")
        ws.cell(row=4, column=36, value=450).font = green_font

        # Apply green background fill to date cell (Col C) for student 2 (marks free enrollment)
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        ws.cell(row=4, column=3).fill = green_fill

        # Row 5: student 3 - Nozanin Jamshedova, date: 29 iyun, total: 0 (current paid), note: 450 (green/overpaid), phone: 99 123 45 67
        ws.cell(row=5, column=1, value=3)
        ws.cell(row=5, column=2, value="Nozanin Jamshedova")
        ws.cell(row=5, column=3, value="29 iyun")
        ws.cell(row=5, column=35, value=0)
        ws.cell(row=5, column=37, value="99 123 45 67")
        ws.cell(row=5, column=36, value=450).font = green_font

        # Row 6: ОБЩИЙ ИТОГ (stops loop)
        ws.cell(row=6, column=2, value="ОБЩИЙ ИТОГ")

        # Save workbook to memory
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        uploaded_file = SimpleUploadedFile(
            name="Kadir_K.xlsx",
            content=excel_buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        self.client.force_authenticate(user=self.admin)
        response = self.client.post('/api/groups/import-excel/', {
            'file': uploaded_file,
            'month': 7,
            'year': 2026,
            'price': 450000.00
        }, format='multipart')
        self.assertEqual(response.status_code, 200)

        # Check response details
        self.assertEqual(response.data['status'], 'success')
        self.assertIn('Kadir K - Tues 900', response.data['imported_groups'])

        # Verify database objects created
        # 1. Teacher User
        teacher = User.objects.get(username="kadir_k")
        self.assertEqual(teacher.first_name, "Kadir K")
        self.assertEqual(teacher.role, "teacher")

        # 2. Group
        group = Group.objects.get(name="Kadir K - Tues 900")
        self.assertEqual(group.teacher, teacher)
        self.assertEqual(group.group_days_at, "Tue-Thur-Sat")
        self.assertEqual(group.starts_at.hour, 9)

        # 3. Students
        s1 = Student.objects.get(full_name="Amirshox Davronov", phone1="+998937277007")
        s2 = Student.objects.get(full_name="Timur Mustafoev", phone1="+998882871772")
        s3 = Student.objects.get(full_name="Nozanin Jamshedova", phone1="+998991234567")

        # 4. Enrollments
        e1 = Enrollment.objects.get(student=s1, group=group)
        self.assertEqual(e1.date.month, 6)
        self.assertEqual(e1.date.day, 24)
        self.assertEqual(e1.status, "dropped") # Excluded due to yellow cell fill
        self.assertFalse(e1.enrolled_free)
        self.assertTrue(e1.pdf_uploaded)

        e2 = Enrollment.objects.get(student=s2, group=group)
        self.assertEqual(e2.date.month, 6)
        self.assertEqual(e2.date.day, 16)
        self.assertEqual(e2.status, "enrolled")
        self.assertTrue(e2.enrolled_free)
        self.assertTrue(e2.pdf_uploaded)

        e3 = Enrollment.objects.get(student=s3, group=group)
        self.assertEqual(e3.date.month, 6)
        self.assertEqual(e3.date.day, 29)
        self.assertEqual(e3.status, "enrolled")
        self.assertFalse(e3.enrolled_free)
        self.assertTrue(e3.pdf_uploaded)

        # 5. Payments
        # Student 1: has 1 payment (current month = 450,000 UZS)
        p1 = Payment.objects.filter(student=s1, group=group, amount=450000.00).count()
        self.assertEqual(p1, 1)

        # Verify enrollment debt details for Student 1
        # Enrolled June 24, today is July 8.
        # Since pdf_uploaded is True: only the current cycle (June 24 - July 24) is evaluated.
        # And they paid 450,000 UZS. So debt should be 0, status 'paid'!
        e1.refresh_from_db()
        self.assertEqual(e1.payment_status, 'paid')
        self.assertEqual(float(e1.debt_amount), 0.00)

        # Student 2 is enrolled free, so no prior cash payments were generated
        p2 = Payment.objects.filter(student=s2, group=group).count()
        self.assertEqual(p2, 0)

        # Student 3 has no current month payment, but has next month payment
        p3_all = Payment.objects.filter(student=s3, group=group).order_by('payment_date')
        self.assertEqual(p3_all.count(), 1)
        self.assertEqual(p3_all[0].amount, 450000.00)
        # Verify payment date is August 28
        self.assertEqual(p3_all[0].payment_date.month, 8)
        self.assertEqual(p3_all[0].payment_date.day, 28)

        # Verify student 3 debt status. Since July is the current month, and they have 0 payments for July:
        e3.refresh_from_db()
        self.assertEqual(e3.payment_status, 'debt')
        self.assertEqual(float(e3.debt_amount), 450000.00)

        # 6. Absences
        absences = Absence.objects.filter(student=s1, group=group)
        self.assertEqual(absences.count(), 1)
        self.assertEqual(absences.first().date.day, 1)

    def test_admin_branch_filtering(self):
        # Create Branch B
        branch_b = Branch.objects.create(name="Branch B")

        # Create Admin user linked to Branch A (self.branch)
        admin_a = User.objects.create_user(
            username="admin_a",
            first_name="Admin A",
            password="adminpassword",
            role="admin",
            branch=self.branch
        )

        # Create Group A (Branch A) and Group B (Branch B)
        group_a = Group.objects.create(
            name="Group A",
            starts_at="10:00:00",
            duration=90,
            started_at=timezone.localdate(),
            branch=self.branch,
            status="ongoing"
        )
        group_b = Group.objects.create(
            name="Group B",
            starts_at="10:00:00",
            duration=90,
            started_at=timezone.localdate(),
            branch=branch_b,
            status="ongoing"
        )

        # Create Teachers
        teacher_a = User.objects.create_user(
            username="teacher_a",
            first_name="Teacher A",
            password="pwd",
            role="teacher",
            branch=self.branch
        )
        teacher_b = User.objects.create_user(
            username="teacher_b",
            first_name="Teacher B",
            password="pwd",
            role="teacher",
            branch=branch_b
        )

        # Create Students
        student_a = Student.objects.create(full_name="Student A", phone1="998900000001")
        student_b = Student.objects.create(full_name="Student B", phone1="998900000002")

        # Enrollments
        Enrollment.objects.create(student=student_a, group=group_a)
        Enrollment.objects.create(student=student_b, group=group_b)

        # Payments
        payment_a = Payment.objects.create(
            group=group_a,
            student=student_a,
            amount=Decimal("10000.00"),
            payment_method="cash",
            status="accepted"
        )
        payment_b = Payment.objects.create(
            group=group_b,
            student=student_b,
            amount=Decimal("20000.00"),
            payment_method="cash",
            status="accepted"
        )

        # Authenticate as admin_a (Branch A admin)
        self.client.force_authenticate(user=admin_a)

        # 1. Verify Group filtering
        response = self.client.get('/api/groups/')
        self.assertEqual(response.status_code, 200)
        group_names = [g['name'] for g in response.data]
        self.assertIn("Group A", group_names)
        self.assertNotIn("Group B", group_names)

        # 2. Verify Teacher/User filtering
        response = self.client.get('/api/users/')
        self.assertEqual(response.status_code, 200)
        teacher_names = [u['first_name'] for u in response.data]
        self.assertIn("Teacher A", teacher_names)
        self.assertNotIn("Teacher B", teacher_names)

        # 3. Verify Student filtering
        response = self.client.get('/api/students/')
        self.assertEqual(response.status_code, 200)
        student_names = [s['full_name'] for s in response.data]
        self.assertIn("Student A", student_names)
        self.assertNotIn("Student B", student_names)

        # 4. Verify Payment filtering
        response = self.client.get('/api/payments/')
        self.assertEqual(response.status_code, 200)
        payment_amounts = [float(p['amount']) for p in response.data]
        self.assertIn(10000.00, payment_amounts)
        self.assertNotIn(20000.00, payment_amounts)

class EndToEndBackendTests(APITestCase):
    def setUp(self):
        # Create a branch, admin user
        self.branch = Branch.objects.create(name="E2E Branch")
        self.admin = User.objects.create_superuser(
            username="e2e_admin",
            password="adminpassword",
            role="superuser"
        )
        self.client.force_authenticate(user=self.admin)

    def test_e2e_crud_operations(self):
        # 1. Course CRUD
        response = self.client.post('/api/courses/', {
            'name': 'New Course',
            'price': '120000.00',
            'description': 'E2E Test Course'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        course_id = response.data['id']

        response = self.client.get('/api/courses/')
        self.assertEqual(response.status_code, 200)
        self.assertTrue(any(c['id'] == course_id for c in response.data))

        response = self.client.patch(f'/api/courses/{course_id}/', {
            'name': 'Updated Course'
        }, format='json')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['name'], 'Updated Course')

        response = self.client.delete(f'/api/courses/{course_id}/')
        self.assertEqual(response.status_code, 204)
        # Verify soft delete
        self.assertFalse(Course.objects.get(id=course_id).is_active)

        # 2. Branch CRUD
        response = self.client.post('/api/branches/', {
            'name': 'New Branch',
            'description': 'E2E Branch desc'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        branch_id = response.data['id']

        response = self.client.get('/api/branches/')
        self.assertEqual(response.status_code, 200)

        response = self.client.patch(f'/api/branches/{branch_id}/', {
            'name': 'Updated Branch'
        }, format='json')
        self.assertEqual(response.status_code, 200)

        # Since BranchViewSet overrides perform_destroy via SoftDeleteModelViewSet, verify soft delete
        response = self.client.delete(f'/api/branches/{branch_id}/')
        self.assertEqual(response.status_code, 204)
        self.assertFalse(Branch.objects.get(id=branch_id).is_active)

        # 3. Room CRUD
        response = self.client.post('/api/rooms/', {
            'name': 'New Room',
            'branch': self.branch.id,
            'description': 'E2E Room desc'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        room_id = response.data['id']

        response = self.client.get('/api/rooms/')
        self.assertEqual(response.status_code, 200)

        response = self.client.patch(f'/api/rooms/{room_id}/', {
            'name': 'Updated Room'
        }, format='json')
        self.assertEqual(response.status_code, 200)

        response = self.client.delete(f'/api/rooms/{room_id}/')
        self.assertEqual(response.status_code, 204)
        self.assertFalse(Room.objects.get(id=room_id).is_active)

        # 4. Student CRUD
        response = self.client.post('/api/students/', {
            'full_name': 'E2E Student',
            'phone1': '998911112233'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        student_id = response.data['id']

        response = self.client.get('/api/students/')
        self.assertEqual(response.status_code, 200)

        response = self.client.patch(f'/api/students/{student_id}/', {
            'full_name': 'Updated Student'
        }, format='json')
        self.assertEqual(response.status_code, 200)

        response = self.client.delete(f'/api/students/{student_id}/')
        self.assertEqual(response.status_code, 204)
        self.assertFalse(Student.objects.get(id=student_id).is_active)

        # 5. User/Teacher CRUD
        response = self.client.post('/api/users/', {
            'username': 'e2e_teacher',
            'first_name': 'E2E Teacher',
            'role': 'teacher',
            'password': 'teacherpassword'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        teacher_id = response.data['id']

        response = self.client.get('/api/users/')
        self.assertEqual(response.status_code, 200)

        response = self.client.patch(f'/api/users/{teacher_id}/', {
            'first_name': 'Updated Teacher'
        }, format='json')
        self.assertEqual(response.status_code, 200)

        response = self.client.delete(f'/api/users/{teacher_id}/')
        self.assertEqual(response.status_code, 204)
        self.assertFalse(User.objects.get(id=teacher_id).is_active)

    def test_e2e_enrollment_payouts_and_debts(self):

        # Setup active student, course, room, group, teacher
        teacher = User.objects.create_user(
            username="active_teacher",
            first_name="Active Teacher",
            role="teacher",
            branch=self.branch
        )
        student = Student.objects.create(full_name="Active Student", phone1="998901234567")
        course = Course.objects.create(name="Active Course", price=Decimal('200000.00'))
        room = Room.objects.create(name="Active Room", branch=self.branch)
        group = Group.objects.create(
            name="Active Group",
            course=course,
            teacher=teacher,
            room=room,
            branch=self.branch,
            price=Decimal('200000.00'),
            starts_at="10:00:00",
            duration=90,
            started_at=timezone.localdate(),
            status="ongoing",
            teacher_share=40
        )

        # 1. Enroll Student via API
        response = self.client.post('/api/enrollments/', {
            'student': student.id,
            'group': group.id,
            'status': 'enrolled'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        enrollment_id = response.data['id']

        # Verify initial debt is set to the group price
        enrollment = Enrollment.objects.get(id=enrollment_id)
        self.assertEqual(float(enrollment.debt_amount), 200000.00)

        # 2. Accept Student Payment via API
        response = self.client.post('/api/payments/', {
            'student': student.id,
            'group': group.id,
            'amount': '200000.00',
            'payment_method': 'cash',
            'status': 'accepted'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        payment_id = response.data['id']

        # Verify debt is now 0 and status is paid
        enrollment.refresh_from_db()
        self.assertEqual(float(enrollment.debt_amount), 0.00)
        self.assertEqual(enrollment.payment_status, 'paid')

        # 3. Pay Teacher (Teacher Payout) via API
        # E.g. teacher earnings should be 200,000 * 40% = 80,000
        # Let's pay them 50,000 UZS
        response = self.client.post('/api/payments/', {
            'student': None,
            'teacher': teacher.id,
            'group': group.id,
            'amount': '50000.00',
            'payment_method': 'cash'
        }, format='json')
        self.assertEqual(response.status_code, 201)
        teacher_payment_id = response.data['id']

        # Authenticate as teacher to confirm receipt
        self.client.force_authenticate(user=teacher)
        confirm_resp = self.client.post(f'/api/payments/{teacher_payment_id}/confirm/')
        self.assertEqual(confirm_resp.status_code, 200)

        # Authenticate back as Admin
        self.client.force_authenticate(user=self.admin)

        # Retrieve group details to verify teacher remaining payout calculation
        response = self.client.get(f'/api/groups/{group.id}/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['teacher_earnings'], 80000.00)
        self.assertEqual(response.data['teacher_paid'], 50000.00)
        self.assertEqual(response.data['teacher_remaining'], 30000.00)

        # 4. Cancel student payment via API (soft delete)
        response = self.client.delete(f'/api/payments/{payment_id}/')
        self.assertEqual(response.status_code, 204)
        
        # Verify payment is soft deleted / canceled
        payment = Payment.objects.get(id=payment_id)
        self.assertFalse(payment.is_active)
        self.assertEqual(payment.status, 'canceled')

        # 5. Call check_student_debts management command
        # After canceling student payment, debt should return to 200,000.
        # Running management command should recalculate this correctly
        call_command('check_student_debts')
        enrollment.refresh_from_db()
        self.assertEqual(float(enrollment.debt_amount), 200000.00)
        self.assertEqual(enrollment.payment_status, 'debt')

class NotificationsAndPayoutConfirmationsTest(APITestCase):
    def setUp(self):
        self.branch = Branch.objects.create(name="Tashkent Branch")
        self.admin = User.objects.create_user(
            username="admin_user",
            password="testpassword",
            role="admin",
            branch=self.branch
        )
        self.teacher = User.objects.create_user(
            username="teacher_user",
            password="testpassword",
            role="teacher"
        )
        self.student = Student.objects.create(full_name="Abubakir Karimov", phone1="998901234567")
        self.group = Group.objects.create(
            name="Intermediate English",
            teacher=self.teacher,
            branch=self.branch,
            started_at=timezone.localdate(),
            starts_at="18:00:00",
            duration=90,
            price=200000.00
        )
        self.enrollment = Enrollment.objects.create(
            student=self.student,
            group=self.group
        )

    def test_absence_creates_notification_for_admin(self):
        self.client.force_authenticate(user=self.teacher)
        data = {
            'student': self.student.id,
            'group': self.group.id,
            'teacher': self.teacher.id,
            'date': str(timezone.localdate())
        }
        response = self.client.post('/api/absences/', data)
        self.assertEqual(response.status_code, 201)

        notifs = Notification.objects.filter(recipient=self.admin, notification_type='absence')
        self.assertEqual(notifs.count(), 1)
        self.assertIn("marked absent", notifs.first().message)

    def test_teacher_payout_confirmation_flow(self):
        self.client.force_authenticate(user=self.admin)
        data = {
            'group': self.group.id,
            'teacher': self.teacher.id,
            'amount': 150000.00,
            'payment_method': 'cash'
        }
        response = self.client.post('/api/payments/', data)
        self.assertEqual(response.status_code, 201)
        payment_id = response.data['id']

        payment = Payment.objects.get(id=payment_id)
        self.assertEqual(payment.status, 'pending')

        notifs = Notification.objects.filter(recipient=self.teacher, notification_type='payment_pending')
        self.assertEqual(notifs.count(), 1)
        self.assertIn("Payout Pending Confirmation", notifs.first().title)

        self.client.force_authenticate(user=self.teacher)
        response = self.client.post(f'/api/payments/{payment_id}/confirm/')
        self.assertEqual(response.status_code, 200)

        payment.refresh_from_db()
        self.assertEqual(payment.status, 'accepted')

        admin_notifs = Notification.objects.filter(recipient=self.admin, notification_type='payment_accepted')
        self.assertEqual(admin_notifs.count(), 1)
        self.assertIn("confirmed receipt", admin_notifs.first().message)
