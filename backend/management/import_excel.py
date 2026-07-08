import os
import re
import datetime
from decimal import Decimal
import openpyxl
from django.utils import timezone
from django.db import transaction
from management.models import User, Student, Group, Enrollment, Payment, Absence, Branch
from management.helpers import is_yellow_color, is_green_color, is_red_color, clean_phone_number, parse_enrollment_date

def run_import_excel(excel_file, month, year, price):
    current_month = int(month)
    current_year = int(year)
    group_price = Decimal(str(price))
    
    # 1. Parse teacher name from file name
    filename = excel_file.name
    base_name = os.path.splitext(filename)[0]
    # Clean name (replace underscores/dashes with spaces)
    teacher_name = base_name.replace('_', ' ').replace('-', ' ').strip().title()
    username = base_name.lower().replace(' ', '_').replace('-', '_')
    
    branch = Branch.objects.first()
    
    # Create or get teacher
    teacher, _ = User.objects.get_or_create(
        username=username,
        defaults={
            'first_name': teacher_name,
            'role': 'teacher',
            'branch': branch
        }
    )
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    
    imported_groups = []
    total_students_imported = 0
    total_payments_imported = 0
    total_absences_imported = 0
    
    with transaction.atomic():
        for sheet in wb.worksheets:
            sheet_name = sheet.title
            
            # Determine group schedule and starts_at from sheet name
            # E.g. "Tues 900" or "Mon 1030"
            sheet_lower = sheet_name.lower()
            if 'tue' in sheet_lower:
                group_days_at = 'Tue-Thur-Sat'
            else:
                group_days_at = 'Mon-Wed-Fri'
                
            # Extract digits for start time
            digits = "".join(c for c in sheet_name if c.isdigit())
            if digits == "900" or digits == "9":
                starts_at = datetime.time(9, 0)
            elif digits == "1030":
                starts_at = datetime.time(10, 30)
            elif len(digits) >= 3:
                try:
                    h = int(digits[:-2])
                    m = int(digits[-2:])
                    starts_at = datetime.time(h, m)
                except ValueError:
                    starts_at = datetime.time(9, 0)
            else:
                starts_at = datetime.time(9, 0)
                
            group_name = f"{teacher_name} - {sheet_name}"
            
            # Create or get group
            group, _ = Group.objects.get_or_create(
                name=group_name,
                defaults={
                    'teacher': teacher,
                    'branch': branch,
                    'price': group_price,
                    'starts_at': starts_at,
                    'duration': 90,
                    'started_at': timezone.localdate(),
                    'group_days_at': group_days_at,
                    'status': 'ongoing'
                }
            )
            imported_groups.append(group_name)
            
            # Parse student records starting at row 3
            row_idx = 3
            while True:
                student_name_val = sheet.cell(row=row_idx, column=2).value
                if not student_name_val or str(student_name_val).strip() == "" or "ОБЩИЙ ИТОГ" in str(student_name_val).upper():
                    break
                    
                student_name = str(student_name_val).strip()
                
                # Column AK (37): Phone number
                phone_raw = sheet.cell(row=row_idx, column=37).value
                phone1 = clean_phone_number(phone_raw)
                
                # Get or create student
                student, _ = Student.objects.get_or_create(
                    full_name=student_name,
                    phone1=phone1
                )
                total_students_imported += 1
                
                # Check background colors of B and C cells for free and dropped
                b_cell = sheet.cell(row=row_idx, column=2)
                c_cell = sheet.cell(row=row_idx, column=3)
                
                b_color = None
                if b_cell.fill and b_cell.fill.start_color and b_cell.fill.start_color.rgb:
                    b_color = b_cell.fill.start_color.rgb
                
                c_color = None
                if c_cell.fill and c_cell.fill.start_color and c_cell.fill.start_color.rgb:
                    c_color = c_cell.fill.start_color.rgb
                    
                is_free = is_green_color(b_color) or is_green_color(c_color)
                is_dropped = is_yellow_color(b_color) or is_yellow_color(c_color)
                
                status = 'dropped' if is_dropped else 'enrolled'
                
                # Column C (3): Enrollment date
                enroll_date_raw = sheet.cell(row=row_idx, column=3).value
                enrollment_date = parse_enrollment_date(enroll_date_raw, current_year, current_month)
                
                # Get or create enrollment (pdf_uploaded is True)
                enrollment, created = Enrollment.objects.get_or_create(
                    student=student,
                    group=group,
                    defaults={
                        'date': enrollment_date,
                        'status': status,
                        'enrolled_free': is_free,
                        'pdf_uploaded': True
                    }
                )
                if not created:
                    enrollment.date = enrollment_date
                    enrollment.status = status
                    enrollment.enrolled_free = is_free
                    enrollment.pdf_uploaded = True
                    enrollment.save()
                    
                # From col D till col AH (4 to 34): Absences if value is "-"
                for col_idx in range(4, 35):
                    day = col_idx - 3
                    cell_val = sheet.cell(row=row_idx, column=col_idx).value
                    if cell_val and str(cell_val).strip() == "-":
                        try:
                            absence_date = datetime.date(current_year, current_month, day)
                            Absence.objects.get_or_create(
                                student=student,
                                group=group,
                                teacher=group.teacher,
                                date=absence_date
                            )
                            total_absences_imported += 1
                        except ValueError:
                            pass
                            
                # Column AI (35): Current month payment
                col_ai_val = sheet.cell(row=row_idx, column=35).value
                if col_ai_val is not None:
                    try:
                        amount = float(str(col_ai_val).replace(' ', '').strip())
                        if amount > 0 and not is_free:
                            if amount < 1000:
                                amount *= 1000
                            Payment.objects.create(
                                group=group,
                                student=student,
                                amount=Decimal(str(amount)),
                                payment_method='cash',
                                status='accepted',
                                description="Imported current month payment from Excel"
                            )
                            total_payments_imported += 1
                    except ValueError:
                        pass
                        
                # Column AJ (36): Next month payment if text/fill is green
                note_cell = sheet.cell(row=row_idx, column=36)
                color_hex = None
                if note_cell.font and note_cell.font.color and note_cell.font.color.rgb:
                    color_hex = note_cell.font.color.rgb
                if not color_hex and note_cell.fill and note_cell.fill.start_color and note_cell.fill.start_color.rgb:
                    color_hex = note_cell.fill.start_color.rgb
                    
                is_green = is_green_color(color_hex)
                if is_green and not is_free:
                    overpaid_val = 0
                    if note_cell.value is not None:
                        try:
                            overpaid_val = float(str(note_cell.value).replace(' ', '').strip())
                            if overpaid_val < 1000:
                                overpaid_val *= 1000
                        except ValueError:
                            pass
                            
                    next_month = current_month + 1
                    next_year = current_year
                    if next_month > 12:
                        next_month = 1
                        next_year += 1
                    
                    next_billing_day = min(enrollment_date.day, 28)
                    next_billing_date = datetime.date(next_year, next_month, next_billing_day)
                    next_pay_amount = overpaid_val if overpaid_val > 0 else float(group_price)
                    
                    Payment.objects.create(
                        group=group,
                        student=student,
                        amount=Decimal(str(next_pay_amount)),
                        payment_method='cash',
                        status='accepted',
                        description="Imported next month payment from Excel",
                        payment_date=timezone.make_aware(datetime.datetime.combine(next_billing_date, datetime.time(12, 0)))
                    )
                    total_payments_imported += 1
                    
                # Recalculate enrollment debt
                enrollment.check_debt()
                
                row_idx += 1
                
    return {
        'status': 'success',
        'imported_groups': imported_groups,
        'total_students': total_students_imported,
        'total_payments': total_payments_imported,
        'total_absences': total_absences_imported
    }
